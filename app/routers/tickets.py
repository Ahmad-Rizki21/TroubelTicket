# app/routers/tickets.py

from fastapi import APIRouter, Depends, HTTPException, Response, status
from fastapi.responses import StreamingResponse
import io
from datetime import datetime

# PDF Generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from typing import List

from .. import schemas, crud, models
from .auth import get_db, has_permission # Import from auth

router = APIRouter(
    prefix="/tickets",
    tags=["Tickets"]
)

@router.post("/", response_model=schemas.Ticket)
def create_ticket(ticket: schemas.TicketCreate, db: Session = Depends(get_db)):
    # Memanggil fungsi create_ticket dari crud.py
    return crud.create_ticket(db=db, ticket=ticket)

@router.get("/", response_model=List[schemas.Ticket])
def read_tickets(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    # Memanggil fungsi get_tickets dari crud.py
    tickets = crud.get_tickets(db, skip=skip, limit=limit)
    return tickets

@router.get("/{ticket_id}", response_model=schemas.Ticket)
def read_ticket(ticket_id: int, db: Session = Depends(get_db)):
    # Memanggil fungsi get_ticket dari crud.py
    db_ticket = crud.get_ticket(db, ticket_id=ticket_id)
    if db_ticket is None:
        # Jika tiket tidak ditemukan, kirim error 404
        raise HTTPException(status_code=404, detail="Ticket not found")
    return db_ticket

@router.put("/{ticket_id}", response_model=schemas.Ticket)
def update_ticket(ticket_id: int, ticket: schemas.TicketUpdate, db: Session = Depends(get_db)):
    db_ticket = crud.update_ticket(db=db, ticket_id=ticket_id, ticket=ticket)
    if db_ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return db_ticket

@router.delete("/{ticket_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_ticket(ticket_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(has_permission("ticket:delete"))):
    # Ambil data tiket terlebih dahulu untuk pengecekan status
    db_ticket = crud.get_ticket(db, ticket_id=ticket_id)
    if db_ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    # Validasi bisnis: Tiket dengan status "Open" tidak bisa dihapus
    if db_ticket.status == "Open":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tiket dengan status 'Open' tidak dapat dihapus. Silakan tutup tiket terlebih dahulu sebelum menghapus."
        )

    # Jika status bukan "Open", lanjutkan proses penghapusan
    deleted_ticket = crud.delete_ticket(db=db, ticket_id=ticket_id)
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# Helper function to calculate duration
def calculate_downtime(start, end):
    if not start:
        return "N/A - No start time"
    
    if not end:
        return "Open - No end time"
    
    # Ensure start and end are datetime objects
    start_dt = start if isinstance(start, datetime) else datetime.fromisoformat(str(start))
    end_dt = end if isinstance(end, datetime) else datetime.fromisoformat(str(end))
    
    # Check if end time is before start time, which indicates a data issue
    if end_dt < start_dt:
        return "Error: Closed before created"
    
    duration = end_dt - start_dt
    
    days = duration.days
    hours, remainder = divmod(duration.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    return f"{days}d {hours}h {minutes}m {seconds}s"

@router.get("/{ticket_id}/export/pdf", tags=["Export"])
def export_ticket_pdf(ticket_id: int, db: Session = Depends(get_db)):
    db_ticket = crud.get_ticket(db, ticket_id=ticket_id)
    if db_ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Ticket Report: {db_ticket.ticket_code}", styles['h1']))
    story.append(Spacer(1, 24))

    # Ticket Details
    story.append(Paragraph("Ticket Details", styles['h2']))
    details_data = [
        ['ID', f': {db_ticket.id}'],
        ['Title', f': {db_ticket.title}'],
        ['Status', f': {db_ticket.status}'],
        ['Priority', f': {db_ticket.priority}'],
        ['Reporter', f': {db_ticket.reporter_name}'],
        ['Created At', f': {db_ticket.created_at.strftime("%Y-%m-%d %H:%M:%S")}'],
        ['Closed At', f': {db_ticket.closed_at.strftime("%Y-%m-%d %H:%M:%S") if db_ticket.closed_at else ": Not Closed"}'],
    ]
    details_table = Table(details_data, colWidths=[100, 350])
    details_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(details_table)
    story.append(Spacer(1, 18))

    # Downtime
    story.append(Paragraph("Total Downtime", styles['h2']))
    downtime = calculate_downtime(db_ticket.created_at, db_ticket.closed_at)
    story.append(Paragraph(downtime, styles['Normal']))
    story.append(Spacer(1, 18))

    # Description
    story.append(Paragraph("Description", styles['h2']))
    story.append(Paragraph(db_ticket.description or '-', styles['Normal']))
    story.append(Spacer(1, 18))

    # Action History
    story.append(Paragraph("Action History", styles['h2']))
    actions = crud.get_actions_for_ticket(db, ticket_id=ticket_id)
    if not actions:
        story.append(Paragraph("No actions recorded for this ticket.", styles['Normal']))
    else:
        action_data = [['Timestamp', 'User', 'Description']]
        for action in actions:
            action_data.append([
                action.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                action.user.username,
                Paragraph(action.action_description, styles['Normal'])
            ])
        action_table = Table(action_data, colWidths=[120, 80, 280])
        action_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(action_table)

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=ticket_{db_ticket.ticket_code}.pdf"
    })

@router.get("/{ticket_id}/export/excel", tags=["Export"])
def export_ticket_excel(ticket_id: int, db: Session = Depends(get_db)):
    db_ticket = crud.get_ticket(db, ticket_id=ticket_id)
    if db_ticket is None:
        raise HTTPException(status_code=404, detail="Ticket not found")

    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Ticket Summary"

    ws_summary.append(['Ticket Summary'])
    ws_summary['A1'].font = Font(bold=True, size=16)
    
    summary_data = {
        "Ticket ID": db_ticket.ticket_code,
        "Title": db_ticket.title,
        "Status": db_ticket.status,
        "Priority": db_ticket.priority,
        "Reporter": db_ticket.reporter_name,
        "Created At": db_ticket.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "Closed At": db_ticket.closed_at.strftime("%Y-%m-%d %H:%M:%S") if db_ticket.closed_at else "Not Closed",
        "Total Downtime": calculate_downtime(db_ticket.created_at, db_ticket.closed_at)
    }

    for key, value in summary_data.items():
        ws_summary.append([key, value])

    # Action History Sheet
    ws_actions = wb.create_sheet(title="Action History")
    actions = crud.get_actions_for_ticket(db, ticket_id=ticket_id)
    
    headers = ['Timestamp', 'User', 'Role', 'Description']
    ws_actions.append(headers)
    for cell in ws_actions[1]:
        cell.font = Font(bold=True)

    for action in actions:
        ws_actions.append([
            action.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            action.user.username,
            action.role.name if action.role else '',
            action.action_description
        ])

    # Save to a memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename=ticket_{db_ticket.ticket_code}.xlsx"
    })


def calculate_downtime(start, end):
    if not start:
        return "N/A - No start time"
    
    if not end:
        return "Open - No end time"
    
    # Ensure start and end are datetime objects
    start_dt = start if isinstance(start, datetime) else datetime.fromisoformat(str(start))
    end_dt = end if isinstance(end, datetime) else datetime.fromisoformat(str(end))
    
    # Check if end time is before start time, which indicates a data issue
    if end_dt < start_dt:
        return "Error: Closed before created"
    
    duration = end_dt - start_dt
    
    days = duration.days
    hours, remainder = divmod(duration.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    return f"{days}d {hours}h {minutes}m {seconds}s"

@router.get("/export/excel", tags=["Export"])
def export_all_tickets_excel(
    db: Session = Depends(get_db),
    showClosed: bool = None,
    dateFrom: str = None,
    dateTo: str = None,
    search: str = None
):
    # Get all tickets with optional filters
    all_tickets = crud.get_tickets(db, skip=0, limit=10000)  # Assuming max 10000 tickets for export
    
    # Apply filters if provided  
    filtered_tickets = all_tickets
    
    # Filter by closed status
    if showClosed is not None and not showClosed:
        filtered_tickets = [ticket for ticket in filtered_tickets if ticket.status != 'Closed']
    
    # Filter by date range
    if dateFrom:
        try:
            date_from = datetime.fromisoformat(dateFrom)
            filtered_tickets = [ticket for ticket in filtered_tickets if ticket.created_at >= date_from]
        except ValueError:
            # Handle case where date is in a different format
            date_from = datetime.strptime(dateFrom, '%Y-%m-%d')
            filtered_tickets = [ticket for ticket in filtered_tickets if ticket.created_at >= date_from]
    
    if dateTo:
        try:
            date_to = datetime.fromisoformat(dateTo)
            date_to = date_to.replace(hour=23, minute=59, second=59)  # End of the day
            filtered_tickets = [ticket for ticket in filtered_tickets if ticket.created_at <= date_to]
        except ValueError:
            # Handle case where date is in a different format
            date_to = datetime.strptime(dateTo, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)  # End of the day
            filtered_tickets = [ticket for ticket in filtered_tickets if ticket.created_at <= date_to]
    
    # Filter by search query
    if search:
        search_lower = search.lower()
        filtered_tickets = [ticket for ticket in filtered_tickets 
                          if search_lower in (ticket.title or '').lower() 
                          or search_lower in (ticket.ticket_code or '').lower()]
    
    all_tickets = filtered_tickets
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All Tickets"
    
    # Add headers
    headers = ['ID', 'Ticket Code', 'Title', 'Description', 'Status', 'Priority', 'Reporter Name', 'Reporter Contact', 'Assignee', 'Created At', 'Closed At', 'Downtime', 'Summary Problem', 'Summary Action']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
    
    # Add ticket data
    for ticket in all_tickets:
        assignee_name = ticket.assignee.username if ticket.assignee else 'Unassigned'
        downtime = calculate_downtime(ticket.created_at, ticket.closed_at)
        ws.append([
            ticket.id,
            ticket.ticket_code,
            ticket.title,
            ticket.description,
            ticket.status,
            ticket.priority,
            ticket.reporter_name,
            ticket.reporter_contact,
            assignee_name,
            ticket.created_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.created_at else '',
            ticket.closed_at.strftime("%Y-%m-%d %H:%M:%S") if ticket.closed_at else '',
            downtime,
            ticket.summary_problem,
            ticket.summary_action
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to a memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=all_tickets.xlsx"
    })